"""
Excel File Generator for Sample Sales Data
Run this script to convert CSV to Excel format
"""

import pandas as pd
from pathlib import Path

def create_excel_from_csv():
    """Convert sample_sales_data.csv to Excel format"""
    
    csv_path = Path(__file__).parent / "sample_sales_data.csv"
    excel_path = Path(__file__).parent / "sample_sales_data.xlsx"
    
    try:
        # Read CSV
        df = pd.read_csv(csv_path)
        
        # Create Excel writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sales Data', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sales Data']
            
            # Format headers
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 12
            
            # Format date column
            for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
            
            # Format quantity column
            for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="right")
        
        print(f"✓ Excel file created successfully!")
        print(f"  Location: {excel_path}")
        print(f"  Records: {len(df)} rows")
        print(f"  Date Range: {df['date'].min()} to {df['date'].max()}")
        print(f"  Total Quantity: {df['quantity'].sum():,} units")
        print(f"  Average Daily: {df['quantity'].mean():.0f} units")
        
    except FileNotFoundError:
        print(f"✗ Error: sample_sales_data.csv not found")
        print(f"  Make sure you're running from: {Path(__file__).parent}")
    except Exception as e:
        print(f"✗ Error: {str(e)}")

if __name__ == "__main__":
    create_excel_from_csv()
